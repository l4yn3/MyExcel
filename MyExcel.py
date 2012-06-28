#-*-coding:utf-8-*-
########################################################################################## 
# 整合Excel 2003 和 Excel 2007 的数据处理（读取和写入）
# author  l4yn3 <620796236@qq.com>
# date    2012-06-28 23-09-50
##########################################################################################
import os
import os.path
import xlrd
import xlwt
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

class AbstractExcel:
    '''Excel抽象类，定义子类中必须实现的方法'''
    ename = None
    exists = True
    wb = None
    sh = None
    w_wb = None
    w_sh = None
    
    def __init__(self, excel_name):
        self.ename = excel_name
        self.CheckExists()
    
    def CheckExists(self):
        '''检查excel文件是否存在'''
        if not self.ename:
            self.exists = False
        if not os.path.isfile(self.ename):
            self.exists = False
    def OpenExcel(self):
        '''打开excel文件，读取的时候使用'''
        pass
    
    def GetSheet(self, num):
        '''读取excel时候获取指定的sheet'''
        pass
        
    def GetRowsNum(self):
        '''获取行数'''
        pass
        
    def GetColsNum(self):
        '''获取列数'''
        pass
    
    def GetCell(self, row, column):
        '''获取某个单元格的内容'''
        pass
    
    def CreateExcel(self):
        '''创建一个excel'''
        pass
    
    def CreateSheet(self, sheet_name):
        '''创建excel当中的sheet'''
        pass
    
    def SetCell(self, row, column, content):
        '''设置单元格当中的内容'''
        pass
    
    def Save(self, excel_name):
        '''保存excel'''
        pass
    
class XlsExcel(AbstractExcel):
    '''处理Excel 2003的子类'''
    def __init__(self, excel_name):
        AbstractExcel.__init__(self, excel_name)
        
    def OpenExcel(self):
        if not self.exists:
            return False;
        self.wb = xlrd.open_workbook(self.ename)
        return self
    
    def GetSheet(self, num):
        self.sh = self.wb.sheet_by_index(int(num))
        return self
    
    def GetRowsNum(self):
        return self.sh.nrows
    
    def GetColsNum(self):
        return self.sh.ncols
    
    def GetCell(self, row, column):
        return str(self.sh.cell(row, column).value)
        
    def CreateExcel(self):
        self.w_wb = xlwt.Workbook()
        return self
    
    def CreateSheet(self, sheet_name):
        self.w_sh = self.w_wb.add_sheet(sheet_name)
        return self
        
    def SetCell(self, row, column, content):
        self.w_sh.write(row, column, content)
        return self
    
    def Save(self):
        return self.w_wb.save(self.ename)
            
class XlsxExcel(AbstractExcel):
    '''处理Excel 2007的子类'''
    eb = None    
    
    def __init__(self, excel_name):
        AbstractExcel.__init__(self, excel_name)
    
    def OpenExcel(self):
        if not self.exists:
            return False
        self.wb = load_workbook(self.ename)
        return self
    
    def GetSheet(self, num):
        self.sh = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[int(num)])
        return self
    
    def GetRowsNum(self):
        return self.sh.get_highest_row()
    
    def GetColsNum(self):
        return self.sh.get_highest_col()
    
    def GetCell(self, _row, _column):
        return str(self.sh.cell(row = _row, column = _column).value)
        
    def CreateExcel(self):
        self.eb = Workbook()
        self.w_wb = ExcelWriter(workbook = self.eb)
        return self
    
    def CreateSheet(self, sheet_name):
        self.w_sh = self.eb.worksheets[int(sheet_name)]
        return self
    
    def SetCell(self, _row, _column, content):
        self.w_sh.cell(row = _row, column = _column).value = content
        return self
    
    def Save(self):
        self.w_wb.save(filename = self.ename)
        
class Factory:
    '''工厂方法，在这个方法里实现对不同Excel格式的处理分发'''
    ename = None
    tools = {'xls':'XlsExcel', 'xlsx':'XlsxExcel'}
    def __init__(self, excel_name):
        self.ename = excel_name
    
    def GetExt(self):
        '''获取Excel文件扩展名'''
        point_index = self.ename.rfind(".")
        return self.ename[point_index+1::]
        
    def Get(self):
        '''根据文件扩展名返回处理相应Excel的类的实例'''
        ext = self.GetExt()
        if ext == 'xls':
            return XlsExcel(self.ename)
        elif ext == 'xlsx':
            return XlsxExcel(self.ename)
